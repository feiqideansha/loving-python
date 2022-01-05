#import the excel
import openpyxl
import time
start = time.time()

input_file = "D:\我爱学习\Computational Methods for Logistics\Assignment\Coding\qa194.xlsx"
input_file_profit = "D:\我爱学习\Computational Methods for Logistics\Assignment\Coding\qa194_profit.xlsx"
excel = openpyxl.load_workbook(input_file)
excel_profit = openpyxl.load_workbook(input_file_profit)
sheet = excel.active
sheet_profit = excel_profit.active

scan_excel = True
nodes_number = 0
nodes_names = []
first_column = 2

while scan_excel is True:
    if sheet.cell(row=1,column=first_column).value is not None:
        nodes_number = nodes_number+1
        nodes_names.append(sheet.cell(row=1, column=first_column).value)
        first_column = first_column+1
    else:
        scan_excel = False


# Calculate the distance
distance = []
for i in range(nodes_number):
    distance.append([])

for i in range(nodes_number):
    for j in range(nodes_number):
        distance[i].append(0)
    
for i in range(nodes_number):
    for j in range(i, nodes_number):
        distance[i][j] = sheet.cell(row=i+2, column=j+2).value
        distance[j][i] = sheet.cell(row=i+2, column=j+2).value

#add profit
nodes_profit = []
for i in range(nodes_number):
    nodes_profit.append(sheet_profit.cell(row=i+1, column=2).value)

#NNH
best_path = 999999999    
T = 8000  #optimal 9352
total_distance = 0
repeat_number = 0
total_profit = 0
next_city = 0
first_city = 0
profit_record = []
route_record = []
distance_record = []
city_visited = []

while first_city < nodes_number:
    city_visited.append(first_city)
    
    while repeat_number < nodes_number :
        best_path = 999999999
         
        for i in range(nodes_number): 
            
            if total_distance + distance[city_visited[-1]][i] + distance[first_city][i] <= T:
                
                if i not in city_visited and distance[city_visited[-1]][i] != 0 and distance[city_visited[-1]][i] <= best_path:
                    best_path = distance[city_visited[-1]][i]
                    next_city = i

        city_visited.append(next_city)
        total_distance = total_distance + best_path
        total_profit = total_profit + nodes_profit[i]
            
        if total_distance <= T:
            repeat_number = repeat_number + 1
        else:
            total_distance = total_distance - best_path
            total_profit = total_profit - nodes_profit[i]
            city_visited.remove(city_visited[-1])
            repeat_number = repeat_number + 1
    
    city_visited.append(first_city)
    route_record.append(city_visited)
    profit_record.append(total_profit)
    distance_record.append(total_distance + distance[first_city][next_city])
    city_visited = []
    repeat_number = 0
    total_distance = 0
    total_profit = 0
    next_city = 0
    first_city = first_city + 1
           
end = time.time()
print('Best route record:', route_record[profit_record.index(max(profit_record))])
print('Best profit record:', max(profit_record))
print('Distance record:', distance_record[profit_record.index(max(profit_record))])
print('Running time:', end-start)





                
        

    

